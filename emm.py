from openpyxl import load_workbook

workbook = load_workbook(filename='appLIST.xlsx')
sheet = workbook.active
cell = sheet.cell(row=3,column=2)
print(cell.value)