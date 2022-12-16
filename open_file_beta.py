def old():
    with open("input.txt", "r") as i:
        input_things = i.read()
    ccc = input_things
    import os
    from termcolor import colored
    from colorama import init
    init()
    from win10toast import ToastNotifier
    toaster = ToastNotifier()
    
    
    """
    try:
        if ('微信' in ccc):
            app_dir = ("C:\ProgramData\Microsoft\Windows\Start Menu\Programs\微信\微信.lnk")
            open_app(app_dir)
            print(colored("已打开WeChat", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开微信')
        elif ('qq' in ccc):
            app_dir = ("C:\ProgramData\Microsoft\Windows\Start Menu\Programs\腾讯软件\TIM\TIM.lnk")
            open_app(app_dir)
            print(colored("已打开QQ", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开QQ')
        elif ('cmd' in ccc):
            os.system("cmd")
            print(colored("已打开cmd", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开cmd')
        elif ('tim' in ccc):
            app_dir = ("C:\ProgramData\Microsoft\Windows\Start Menu\Programs\腾讯软件\TIM\TIM.lnk")
            open_app(app_dir)
            print(colored("已打开tim", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开tim')
        elif ('vmware' in ccc):
            app_dir = ("C:\Program Files (x86)\VMware\VMware Workstation\vmware.exe")
            open_app(app_dir)
            print(colored("已打开vmware", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开VMware')
        elif ('google' in ccc):
            app_dir = ("C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Google Chrome.lnk")
            open_app(app_dir)
            print(colored("已打开google", "green", "on_white"))
            toaster.show_toast(u'LittleAI', u'已打开GoogleChrome')

        else:
            print(colored("目前只支持'微信','qq','cmd','tim','vmware','google'  区分大小写!", "red"))
            toaster.show_toast(u"目前只支持'微信','qq','cmd','tim','vmware','google'", u"区分大小写!")
    except FileNotFoundError:
        print("你似乎还没有安装这个软件 -0")

    """



def list_app():
    import os
    import pandas as pd
    from openpyxl import load_workbook
    workbook = load_workbook(filename='appLIST.xlsx')
    sheet = workbook.active
    # 定义函数
    def list_all_files(rootdir):
        import os
        _files = []
        # 列出文件夹下所有的目录与文件
        list = os.listdir(rootdir)
        for i in range(0, len(list)):
            # 构造路径
            path = os.path.join(rootdir, list[i])
            # 判断路径是否为文件目录或者文件
            # 如果是目录则继续递归
            if os.path.isdir(path):
                _files.extend(list_all_files(path))
            if os.path.isfile(path):
                _files.append(path)
        return _files
    # 执行
    dir = r'C:\\ProgramData\\Microsoft\\Windows\\Start Menu\\Programs' # 目录地址
    datatx = pd.DataFrame(list_all_files(dir))
    datatx.to_excel('appLIST.xlsx')
    dfwr = pd.read_excel("appLIST.xlsx", usecols=[1])
    print(dfwr.values)


def openf():
    import os
    import pandas as pd
    from openpyxl import load_workbook
    import pandas
    import pandas as pd
    import numpy as np
    from termcolor import colored
    from colorama import init
    with open("input.txt", "r") as i:
        input_things = i.read()
    f = (input_things[5:])
    def open_app(app_dir):
        os.startfile(app_dir)
    
    """
    找到坐标
    """
    
    def get_coordinates(data: pandas.DataFrame, target: str):
        """
        根据要查找的目标,返回其在excel中的位置
        data: excel数据,
        target: 要查找的目标
        return: 返回坐标列表
        """
        data_list = np.array(data).tolist()
        for i in range(len(data_list)):
            for j in range(len(data_list[i])):
                # print(i, j, data_list[i][j], target, '\n')
                # if str(data_list[i][j]) == str(target):
                if str(target) in str(data_list[i][j]):
                    print("find")
                    return [i + 2, j + 1]
        return []
    data = pd.read_excel('appLIST.xlsx')
    target_string = f
    coordinates = get_coordinates(data, target_string)
    df = pd.read_excel("appLIST.xlsx", usecols=[1])
    
    print(f)
    if str(target_string) in str(df.values):
        print("True. -1")
        # print(f"{target_string}在第{coordinates[0]}行,第{coordinates[1]}列")
        # print(coordinates)
        h = ((str(coordinates[:1]))[:-1])[1:] #去除中括号
        l = 2
    else:
        print(colored("Error! NO software coincidental founded.  -0", "red"))
        print(colored("you can try 'list app' ", "red"))


    """
    打开文件
    """
    workbook = load_workbook(filename='appLIST.xlsx')
    sheet = workbook.active
    cell = sheet.cell(row=int(h),column=int(l))
    print(cell.value)
    open_app(cell.value)

